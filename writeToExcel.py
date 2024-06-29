from openpyxl import Workbook, load_workbook

def write_to_excel(file_path, sheet_name, data):
    try:
        # Load the workbook or create a new one if it doesn't exist
        try:
            workbook = load_workbook(file_path)
        except FileNotFoundError:
            workbook = Workbook()

        # Select the target sheet or create a new one if it doesn't exist
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        # Write data to the sheet
        for row in data:
            sheet.append(row)

        # Save the workbook
        workbook.save(file_path)
        print("Data has been successfully written to the Excel workbook.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
file_path = "example.xlsx"
sheet_name = "plum"
data = [
    ["gah", "goo", "City"],
    ["John", 45, "New York"],
    ["Alice", 100, "Boston"],
    ["Bob", 35, "rim"]
]

write_to_excel(file_path, sheet_name, data)
